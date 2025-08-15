import os
import sys
import openpyxl
from tqdm import tqdm
from docx import Document
import re
import spacy
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from src.config import config
from src.llm_client import llm_client

# 验证LLM配置
provider_info = llm_client.get_provider_info()
if not provider_info['has_api_key']:
    print(f"错误: 未配置{provider_info['provider']} API密钥")
    if config.llm_provider == 'openai':
        print("请在.env文件中设置OPENAI_API_KEY")
    elif config.llm_provider == 'deepseek':
        print("请在.env文件中设置DEEPSEEK_API_KEY")
    sys.exit(1)

print(f"✅ 使用 {provider_info['provider']} - {provider_info['model']}")

# 加载 Spacy 的中文模型，用于句子的分割
try:
    nlp = spacy.load('zh_core_web_sm')
except OSError:
    print("错误: 未安装中文spaCy模型，请运行: python -m spacy download zh_core_web_sm")
    sys.exit(1)


# 定义一个函数，将字符数少于设定值的句子进行合并
def merge_short_sentences(sentences, min_length=None):
    """合并短句子以确保语义完整性"""
    if min_length is None:
        min_length = config.min_sentence_length
    
    merged_sentences = []
    buffer_sentence = ""

    for sentence in sentences:
        sentence = sentence.strip()
        if len(buffer_sentence + sentence) < min_length:
            buffer_sentence += " " + sentence if buffer_sentence else sentence
        else:
            if buffer_sentence:
                merged_sentences.append(buffer_sentence)
            buffer_sentence = sentence

    if buffer_sentence:
        merged_sentences.append(buffer_sentence)

    return merged_sentences


# 定义一个函数，翻译文本为英文
def translate_to_english(text):
    """翻译文本为英文"""
    return llm_client.translate_to_english(text)

# 定义一个函数，将文本翻译为分镜脚本

def translate_to_storyboard(text):
    """将文本翻译为分镜脚本"""
    messages = [
        {"role": "system", "content": "You are a professional storyboard assistant."},
        {"role": "user", "content": f"Based on the text \"{text}\", 基于我给的文本，在提示的生成中，您需要使用提示词来描述角色属性、主题、外表、情感、服装、姿势、视角、行动、背景。使用英语单词或短语甚至自然语言标签进行描述不仅限于我给你的单词。每次生成一组提示词，提示词应该使用英语词组或简短的句子，每组不能大于75个token。然后将您想要的类似提示词组合在一起，使用英语半角作为分隔符，并按从最重要到最不重要的顺序排列。在角色属性中，1girl表示你生成了一个女孩，1boy表示你生成了一个男孩，人数可以是多个。另请注意，提示不能包含-和_。可以有空格和自然语言，但不能太多，单词不能重复。包括角色的性别、主题、外表、情感、服装、姿势、视角、行动、背景，并按从最重要到最不重要的顺序排列，越靠前的词组权重越大；可以使用括号+数字表示增加或减少权重，例如 (depth effect:1.16)，数字的范围是0-2，大于1表示增加权重，小于1表示减少权重，权重一般不超过1.5； 需要优先表现的提示词才可以增加权重，且增加权重的提示词不超过3个；优先级定义：性别>情感>外表>服装>主题>姿势>主题>行动>视角>背景>其它。我会给出一些结构示例 :1boy, club setting, daily routine, viewing others, 1girl, 18 years old, (long red hair:1.5), (casual clothing:1.2), (standing pose:1.1), side view, nightlife background。提示单词应以纯文本输出，不需要有任何解释或示例，不加引号。不要回答不必要的内容，不要问我，也不要给我举例。"},
    ]
    return llm_client.chat_completion(messages)

# 定义一个函数，读取 docx 文件中的段落
def read_docx(file_path):
    document = Document(file_path)
    paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    if not paragraphs:
        raise ValueError("未能读取到有效的文本内容")
    return paragraphs

# 定义一个函数，将分割后的句子写入到 Excel 文件的 A 列
def write_to_excel(sentences, workbook):
    sheet = workbook.active
    for idx, sentence in enumerate(sentences, 1):
        if sentence.strip():  # Ignore blank lines
            sheet.cell(row=idx, column=1, value=sentence)

def replace_text_in_sentences(sentences, original_text, new_text):
    return [sentence.replace(original_text, new_text) for sentence in sentences]

def process_text_sentences(workbook, input_file_path, output_file_path):
    """处理文本句子，生成翻译和分镜脚本"""
    try:
        paragraphs = read_docx(input_file_path)
    except ValueError as e:
        print(f"发生错误：{str(e)}")
        return False

    sentences = []
    for paragraph in paragraphs:
        sentences.extend(re.findall('.*?[。！？]', paragraph))

    sentences = merge_short_sentences(sentences)  

    write_to_excel(sentences, workbook)  # 先将原始句子写入到Excel的A列

    while True:
        input_text = input("请输入要被替换的文字和需要绑定的数字（格式：“原文 新文 数字”，n/N结束）:")
        if input_text.lower() == 'n':
            break
        original_text, new_text, number = input_text.split()
        sentences = replace_text_in_sentences(sentences, original_text, new_text)

        sheet = workbook.active
        for idx, sentence in enumerate(sentences, 1):
            if new_text in sentence:  # 如果新文本在句子中，将数字填入到第5列
                sheet.cell(row=idx, column=5, value=number)
            else:
                if sheet.cell(row=idx, column=5).value is None:  # 如果新文本不在句子中，并且第5列尚未被填充，将0填入到第5列
                    sheet.cell(row=idx, column=5, value=0)

    sheet = workbook.active
    max_workers = min(len(sentences), config.max_workers_translation)            

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures_translation = {executor.submit(translate_to_english, sentence.strip()): idx 
                               for idx, sentence in enumerate(sentences, 1)}
        futures_storyboard = {} 

        for future in tqdm(as_completed(futures_translation), total=len(futures_translation), desc='正在翻译文本'):
            idx = futures_translation[future]
            translated_text = future.result()
            sheet.cell(row=idx, column=2, value=translated_text)
            sheet.cell(row=idx, column=4, value=sentences[idx-1])  # Add replaced sentences to 'D' column
            futures_storyboard[executor.submit(translate_to_storyboard, translated_text)] = idx

        for future in tqdm(as_completed(futures_storyboard), total=len(futures_storyboard), desc='正在生成故事板'):
            idx = futures_storyboard[future]
            sheet.cell(row=idx, column=3, value=future.result())

    workbook.save(output_file_path)
    return True

def main():
    """主函数：执行文本分析和分镜脚本生成"""
    input_file_path = config.input_docx_file
    output_file_path = config.output_excel_file
    
    print(f"Step 1: AI文本分析和分镜脚本生成")
    print(f"输入文件: {input_file_path}")
    print(f"输出文件: {output_file_path}")
    
    # 检查输入文件是否存在
    if not input_file_path.exists():
        print(f"错误: 输入文件不存在 - {input_file_path}")
        return False
    
    # 确保输出目录存在
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 验证配置
    errors = config.validate_config()
    if errors:
        print("配置错误:")
        for error in errors:
            print(f"  - {error}")
        return False
    
    try:
        workbook = openpyxl.Workbook()
        success = process_text_sentences(workbook, input_file_path, output_file_path)
        
        if success:
            print(f"成功生成分析结果: {output_file_path}")
            return True
        else:
            print("处理失败")
            return False
            
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
        return False

if __name__ == '__main__':
    success = main()
    if not success:
        sys.exit(1)